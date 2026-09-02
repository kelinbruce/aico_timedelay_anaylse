"""
分析 nextagent-operational 日志 - knowledge-qa skill 性能 (v3)
适配新日志格式: timestamp带+08:00后缀, RAG事件名变更, RAG请求体结构变更
支持两种评测集: 无线(Col 72) / 云核(Col 71) - 通过命令行参数选择
修复 Workflow 循环执行同名 capability 时，阶段时间线错误配对导致负耗时的问题

用法:
  python analyze_65_kq_v3.py <日志目录> <场景>
  例: python analyze_65_kq_v3.py "D:/temp/20260717/AF2.0性能/71.77.65.24_0805_无线500题" wireless
  例: python analyze_65_kq_v3.py "D:/temp/20260717/AF2.0性能/71.77.65.24_0801晚_云核100题" cloud
"""
import json
import os
import re
import statistics
import argparse
from collections import defaultdict, Counter
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl import load_workbook as lb
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 命令行参数 =====
parser = argparse.ArgumentParser(description='分析 nextagent-operational 日志 - knowledge-qa skill 性能')
parser.add_argument('log_dir', help='日志目录路径')
parser.add_argument('scenario', choices=['wireless', 'cloud'], help='评测场景: wireless(无线) 或 cloud(云核)')
parser.add_argument('--t-start', help='手动指定开始时间(UTC), 格式: YYYY-MM-DD HH:MM:SS, 不指定则自动检测')
parser.add_argument('--t-end', help='手动指定结束时间(UTC), 格式: YYYY-MM-DD HH:MM:SS, 不指定则自动检测')
args = parser.parse_args()

LOG_DIR = args.log_dir
TEST_SET_CONFIG = args.scenario

# 评测集配置
TEST_SETS = {
    'wireless': {
        'path': 'D:/temp/20260717/AF2.0性能/评测集/无线-中文-知识问答-500题-NoeMate2_0_24.xlsx',
        'level_col': 72,  # Col 72: 难度等级 (Level 1/2/3/4)
        'name': '无线',
    },
    'cloud': {
        'path': 'D:/temp/20260717/AF2.0性能/评测集/云核_中文_知识问答_1014题_最新.xlsx',
        'level_col': 71,  # Col 71: 级别 (Level 1/2/3/4/5)
        'name': '云核',
    },
}
TEST_SET_PATH = TEST_SETS[TEST_SET_CONFIG]['path']
LEVEL_COL = TEST_SETS[TEST_SET_CONFIG]['level_col']

def parse_time(t):
    """Parse timestamp supporting both 'Z' and '+08:00' suffixes, returns naive UTC datetime."""
    if not t:
        return None
    try:
        # Handle +08:00 suffix - convert to UTC
        if '+08:00' in t:
            dt_str = t.replace('+08:00', '')
            dt = datetime.strptime(dt_str, '%Y-%m-%dT%H:%M:%S.%f')
            return dt - timedelta(hours=8)
        # Handle Z suffix
        if t.endswith('Z'):
            return datetime.strptime(t, '%Y-%m-%dT%H:%M:%S.%fZ')
        # Try occurredAt format
        if t.endswith('+00:00'):
            dt_str = t.replace('+00:00', '')
            return datetime.strptime(dt_str, '%Y-%m-%dT%H:%M:%S.%f')
        return datetime.strptime(t, '%Y-%m-%dT%H:%M:%S.%fZ')
    except:
        try:
            return datetime.strptime(t[:19], '%Y-%m-%dT%H:%M:%S')
        except:
            return None


def capability_duration_seconds(completed, started=None):
    """Return a non-negative capability duration, preferring the producer's durationMs."""
    duration_ms = completed.get('durationMs')
    if isinstance(duration_ms, (int, float)) and duration_ms >= 0:
        return round(duration_ms / 1000, 3)
    if started and started.get('timestamp') and completed.get('timestamp'):
        duration = (completed['timestamp'] - started['timestamp']).total_seconds()
        if duration >= 0:
            return round(duration, 3)
    return None


def remember_capability_start(by_span, pending_by_invocation, event):
    """Remember a start by unique spanId and by invocation-id FIFO for legacy logs."""
    span_id = event.get('spanId', '')
    if span_id:
        by_span[span_id] = event
    pending_by_invocation[event.get('capabilityInvocationId', '')].append(event)


def pop_capability_start(by_span, pending_by_invocation, completed):
    """Match completed to started by spanId; fall back to invocation-id FIFO."""
    invocation_id = completed.get('capabilityInvocationId', '')
    queue = pending_by_invocation.get(invocation_id, [])
    span_id = completed.get('spanId', '')
    started = by_span.pop(span_id, None) if span_id else None
    if started is not None:
        try:
            queue.remove(started)
        except ValueError:
            pass
        return started
    if queue:
        started = queue.pop(0)
        started_span_id = started.get('spanId', '')
        if started_span_id:
            by_span.pop(started_span_id, None)
        return started
    return None

# ===== 自动检测时间范围 =====
if args.t_start and args.t_end:
    T_START = datetime.strptime(args.t_start, '%Y-%m-%d %H:%M:%S')
    T_END = datetime.strptime(args.t_end, '%Y-%m-%d %H:%M:%S')
    LOCAL_TIME_DESC = f"{(T_START + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')} ~ {(T_END + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')}"
else:
    print("自动检测日志时间范围...")
    all_timestamps = []
    for fname in sorted(os.listdir(LOG_DIR)):
        if not fname.endswith('.jsonl'):
            continue
        fpath = os.path.join(LOG_DIR, fname)
        with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    ts = parse_time(obj.get('timestamp', ''))
                    if ts:
                        all_timestamps.append(ts)
                except:
                    continue
    if all_timestamps:
        T_START = min(all_timestamps)
        T_END = max(all_timestamps)
        LOCAL_TIME_DESC = f"{(T_START + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')} ~ {(T_END + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M')}"
    else:
        raise ValueError("无法从日志中检测到时间戳，请使用 --t-start 和 --t-end 手动指定")
    print(f"  检测到时间范围(UTC): {T_START.strftime('%Y-%m-%d %H:%M:%S')} ~ {T_END.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  本地时间(+0800): {LOCAL_TIME_DESC}")

# ========== 0. Load test set: question -> difficulty level ==========
print("Skip test set loading.")
question_to_level = {}

def match_question(preview, min_prefix=10):
    """Match inputTextPreview to test set question with tolerance, return (matched_question, difficulty_level)."""
    if not preview:
        return None, ''
    preview = preview.strip()
    # Exact
    if preview in question_to_level:
        return preview, question_to_level[preview]
    # Substring
    for q in question_to_level:
        if preview in q or q in preview:
            return q, question_to_level[q]
    # Prefix match
    for q in question_to_level:
        prefix_len = 0
        for i in range(min(len(preview), len(q))):
            if preview[i] == q[i]:
                prefix_len += 1
            else:
                break
        if prefix_len >= min_prefix:
            return q, question_to_level[q]
    return None, ''

# ========== 1. Parse relevant events ==========
RELEVANT_EVENTS = {
    'capability.started', 'capability.completed',
    'model.invocation.started', 'model.invocation.completed',
    'hook.completed',
    'request.accepted', 'request.completed', 'request.failed',
    'policy.allowed', 'policy.denied',
    'runtime.run.dispatched', 'runtime.run.execution_finished',
    'reference-remote-rag-request', 'reference-remote-rag-response',
    'reference-remote-workflow-rag-request', 'workflow_rag_retrieval_completed',
    'session.title.generated',
    'capability.invocation.error', 'capability.failed', 'capability.timed.out',
    'capability.denied', 'capability.canceled',
    'request.input_text',
    'tool.payload.captured',
}

session_events = defaultdict(list)
rag_requests = []   # list of request events
rag_responses = []  # list of response events
session_query = {}  # sessionId -> inputTextPreview (from workflow events)

total_lines = 0
relevant_lines = 0
skipped_time = 0

# Phase 1: Scan ALL files for workflow events (inputTextPreview) regardless of time window
# These events may be in earlier log files but still belong to sessions in our window
print("Phase 1: Scanning for workflow events (inputTextPreview)...")
for fname in sorted(os.listdir(LOG_DIR)):
    if not fname.endswith('.jsonl'):
        continue
    fpath = os.path.join(LOG_DIR, fname)
    with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except:
                continue
            evt = obj.get('event', '')
            sid = obj.get('sessionId', '')
            preview = ''
            if evt == 'workflow.remoteServiceCall.execute.start':
                preview = obj.get('inputTextPreview', '')
            elif evt == 'request.accepted':
                # New path: inputTextPreview in details (from observation pipeline)
                preview = (obj.get('details') or {}).get('inputTextPreview', '')
            elif evt == 'request.input_text':
                # New path: direct diagnosticLogger.info output
                preview = obj.get('inputTextPreview', '')
            if sid and preview:
                session_query[sid] = preview

print(f"  Found inputTextPreview for {len(session_query)} sessions (sources: workflow.remoteServiceCall / request.accepted / request.input_text)")

# Phase 2: Parse relevant events
# Include: (a) sessions within our time window, (b) sessions with inputTextPreview (even if before window)
# We need the timing data for inputTextPreview sessions too, so parse their events
preview_sids_in_logs = set(session_query.keys())

# Select files: all files that could contain events for our window sessions OR preview sessions
files_to_process = []
for fname in sorted(os.listdir(LOG_DIR)):
    if not fname.endswith('.jsonl'):
        continue
    fpath = os.path.join(LOG_DIR, fname)
    with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
        first_line = f.readline().strip()
        while first_line == '' :
            first_line = f.readline().strip()
            if not first_line:
                break
        if not first_line:
            continue
        try:
            first_obj = json.loads(first_line)
            first_ts = parse_time(first_obj.get('timestamp', ''))
            if first_ts and first_ts > T_END:
                continue
        except:
            pass
    files_to_process.append((fname, fpath))

print(f"Files to process: {len(files_to_process)}")

for fname, fpath in files_to_process:
    print(f"  Processing {fname}...")
    with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            total_lines += 1
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except:
                continue

            event = obj.get('event', '')
            if event not in RELEVANT_EVENTS:
                continue

            sid = obj.get('sessionId', '')
            ts = parse_time(obj.get('timestamp', ''))

            # Include event if: session is in our time window, OR session has inputTextPreview
            if sid and sid in preview_sids_in_logs:
                # Keep all events for preview sessions (even outside time window)
                pass
            elif ts and (ts < T_START or ts > T_END):
                skipped_time += 1
                continue

            relevant_lines += 1

            entry = {
                'timestamp': ts,
                'event': event,
                'sessionId': sid,
                'requestId': obj.get('requestId', ''),
                'runId': obj.get('runId', ''),
                'capabilityInvocationId': obj.get('capabilityInvocationId', ''),
                'spanId': obj.get('spanId', ''),
                'details': obj.get('details', {}),
                'durationMs': obj.get('durationMs'),
                'toolCallId': obj.get('toolCallId', ''),
                'toolInput': obj.get('toolInput', {}),
                'toolSafeSummary': obj.get('toolSafeSummary', ''),
                'raw': obj,
            }

            if sid:
                session_events[sid].append(entry)

            # Only count RAG events within time window
            if ts and T_START <= ts <= T_END:
                if event in ('reference-remote-rag-request', 'reference-remote-workflow-rag-request'):
                    rag_requests.append(entry)
                elif event in ('reference-remote-rag-response', 'workflow_rag_retrieval_completed'):
                    rag_responses.append(entry)

print(f"\nTotal lines: {total_lines}")
print(f"Skipped (out of time window): {skipped_time}")
print(f"Relevant events: {relevant_lines}")
print(f"Sessions: {len(session_events)}")
print(f"RAG requests: {len(rag_requests)}")
print(f"RAG responses: {len(rag_responses)}")

# ========== 2. Identify knowledge-qa sessions ==========
# Sessions with Skill capability OR sessions with inputTextPreview
kq_session_ids = set()
for sid, events in session_events.items():
    for e in events:
        if e['event'] == 'capability.started' and e['details'].get('capabilityId') == 'Skill':
            kq_session_ids.add(sid)
            break
    if sid in preview_sids_in_logs and sid not in kq_session_ids:
        kq_session_ids.add(sid)  # Add preview sessions even without Skill events

print(f"Sessions with Skill invocation or inputTextPreview: {len(kq_session_ids)}")

# ========== 3. RAG request-response matching ==========
# Sort requests by timestamp for efficient matching
rag_requests_sorted = sorted(rag_requests, key=lambda e: e['timestamp'] or datetime.min)
rag_responses_sorted = sorted(rag_responses, key=lambda e: e['timestamp'] or datetime.min)

rag_matched = []
used_req_indices = set()

for resp in rag_responses_sorted:
    resp_ts = resp['timestamp']
    if not resp_ts:
        continue
    best_idx = None
    best_diff = None
    for i, req in enumerate(rag_requests_sorted):
        if i in used_req_indices:
            continue
        req_ts = req['timestamp']
        if not req_ts:
            continue
        if req_ts > resp_ts:
            break
        diff = (resp_ts - req_ts).total_seconds()
        if diff < 0 or diff > 300:  # skip if >5min gap
            continue
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_idx = i
    if best_idx is not None:
        used_req_indices.add(best_idx)
        req = rag_requests_sorted[best_idx]
        raw = req['raw']
        # Support both old format (request.indexes) and new format (requestBody.ragIndexes)
        req_obj = raw.get('request', {}) or raw.get('requestBody', {})
        rag_indexes = req_obj.get('ragIndexes', [])
        if isinstance(rag_indexes, list) and len(rag_indexes) > 0 and isinstance(rag_indexes[0], dict):
            indexes = [ri.get('ragIndex', '') for ri in rag_indexes]
        else:
            indexes = req_obj.get('indexes', [])
        query = req_obj.get('query', '')
        topK = req_obj.get('topK', '') or req_obj.get('rankTopN', '')
        rag_matched.append({
            'request_ts': req['timestamp'],
            'response_ts': resp_ts,
            'duration_s': round(best_diff, 3),
            'query': query,
            'indexes': indexes,
            'indexCount': len(indexes),
            'topK': topK,
            'statusCode': resp['raw'].get('statusCode', '') or resp['raw'].get('status', ''),
            'bodyLength': resp['raw'].get('bodyLength', 0),
        })

print(f"RAG matched pairs: {len(rag_matched)}")

# RAG indexes statistics
index_counts = [r['indexCount'] for r in rag_matched]
index_name_counter = Counter()
for r in rag_matched:
    for idx in r['indexes']:
        index_name_counter[idx] += 1

# ========== 4. Build timing analysis per session ==========
all_rows = []

for sid in sorted(kq_session_ids):
    events = sorted(session_events[sid], key=lambda e: e['timestamp'] or datetime.min)

    all_times = [e['timestamp'] for e in events if e['timestamp']]
    if not all_times:
        continue
    session_start = min(all_times)
    session_end = max(all_times)
    session_duration = round((session_end - session_start).total_seconds(), 2)

    # Model invocations - use ordered lists to match started/completed pairs
    model_durations = []
    model_details = []  # (duration, finishReason, resolvedToolNames)
    model_started_pending = {}  # runId -> list of started events (FIFO)
    for e in events:
        if e['event'] == 'model.invocation.started':
            key = e['runId']
            if key not in model_started_pending:
                model_started_pending[key] = []
            model_started_pending[key].append(e)
        elif e['event'] == 'model.invocation.completed':
            key = e['runId']
            start_e = None
            if key in model_started_pending and model_started_pending[key]:
                start_e = model_started_pending[key].pop(0)
            dur = None
            if start_e and start_e['timestamp'] and e['timestamp']:
                dur = round((e['timestamp'] - start_e['timestamp']).total_seconds(), 3)
            elif e.get('durationMs') is not None:
                dur = round(e['durationMs'] / 1000, 3)
            if dur is not None and dur >= 0:
                model_durations.append(dur)
                fr = e['details'].get('finishReason', '') if e['details'] else ''
                tools = e['details'].get('resolvedToolNames', []) if e['details'] else []
                model_details.append((dur, fr, tools))

    model_total = round(sum(model_durations), 2) if model_durations else 0
    model_count = len(model_durations)

    # Capability invocations
    cap_details = defaultdict(lambda: {'total': 0.0, 'count': 0, 'durations': []})
    cap_started_by_span = {}
    cap_started_pending = defaultdict(list)
    # Map capabilityInvocationId -> skill/tool name (from tool.payload.captured)
    cap_inv_names = {}
    for e in events:
        if e['event'] == 'tool.payload.captured':
            # toolCallId in tool.payload = capabilityInvocationId in capability.started/completed
            cap_inv_id = e.get('toolCallId', '') or e.get('capabilityInvocationId', '')
            inp_name = (e.get('toolInput', {}) or {}).get('name', '')
            safe_summary = e.get('toolSafeSummary', '')
            if inp_name:
                cap_inv_names[cap_inv_id] = inp_name
            elif safe_summary:
                cap_inv_names[cap_inv_id] = safe_summary.split()[0] if safe_summary.split() else ''

    for e in events:
        if e['event'] == 'capability.started':
            remember_capability_start(cap_started_by_span, cap_started_pending, e)
        elif e['event'] == 'capability.completed':
            start_e = pop_capability_start(cap_started_by_span, cap_started_pending, e)
            cap_id = (e.get('details') or {}).get('capabilityId', 'Unknown')
            if cap_id == 'Unknown' and start_e:
                cap_id = (start_e.get('details') or {}).get('capabilityId', 'Unknown')
            dur = capability_duration_seconds(e, start_e)
            if dur is not None:
                cap_details[cap_id]['total'] += dur
                cap_details[cap_id]['count'] += 1
                cap_details[cap_id]['durations'].append(dur)

    cap_total = round(sum(v['total'] for v in cap_details.values()), 2)
    cap_count = sum(v['count'] for v in cap_details.values())
    other_duration = round(max(0, session_duration - model_total - cap_total), 2)

    # Skill (knowledge-qa)
    skill_durations = cap_details.get('Skill', {}).get('durations', [])
    skill_count = cap_details.get('Skill', {}).get('count', 0)
    skill_total = round(cap_details.get('Skill', {}).get('total', 0), 2)

    # Rag
    rag_durations_list = cap_details.get('Rag', {}).get('durations', [])
    rag_count = cap_details.get('Rag', {}).get('count', 0)
    rag_total = round(cap_details.get('Rag', {}).get('total', 0), 2)

    # Skill start/end times
    skill_start_times = []
    skill_end_times = []
    for e in events:
        if e['event'] == 'capability.started' and e['details'].get('capabilityId') == 'Skill':
            skill_start_times.append(e['timestamp'].strftime('%Y-%m-%dT%H:%M:%S') if e['timestamp'] else '')
        if e['event'] == 'capability.completed' and e['details'].get('capabilityId') == 'Skill':
            skill_end_times.append(e['timestamp'].strftime('%Y-%m-%dT%H:%M:%S') if e['timestamp'] else '')

    # Stage timeline with step details
    stage_timeline = []
    # Interleave model and capability events in chronological order
    # Rebuild pending start events for the step timeline. A capabilityInvocationId
    # identifies a workflow node and may be reused by multiple loop iterations;
    # spanId uniquely identifies each concrete invocation.
    model_started_pending2 = {}
    cap_started_by_span2 = {}
    cap_started_pending2 = defaultdict(list)
    step_events = []
    for e in events:
        evt = e['event']
        if evt == 'model.invocation.started':
            key = e['runId']
            if key not in model_started_pending2:
                model_started_pending2[key] = []
            model_started_pending2[key].append(e)
        elif evt == 'model.invocation.completed':
            key = e['runId']
            start_e = None
            if key in model_started_pending2 and model_started_pending2[key]:
                start_e = model_started_pending2[key].pop(0)
            if start_e and start_e['timestamp'] and e['timestamp']:
                dur = round((e['timestamp'] - start_e['timestamp']).total_seconds(), 3)
            elif e.get('durationMs') is not None:
                dur = round(e['durationMs'] / 1000, 3)
            else:
                dur = 0
            fr = e['details'].get('finishReason', '') if e['details'] else ''
            tools = e['details'].get('resolvedToolNames', []) if e['details'] else []
            step_events.append((e['timestamp'], 'model', dur, fr, tools, None))
        elif evt == 'capability.started':
            remember_capability_start(cap_started_by_span2, cap_started_pending2, e)
        elif evt == 'capability.completed':
            start_e = pop_capability_start(cap_started_by_span2, cap_started_pending2, e)
            cap_id = (e.get('details') or {}).get('capabilityId', 'Unknown')
            if cap_id == 'Unknown' and start_e:
                cap_id = (start_e.get('details') or {}).get('capabilityId', 'Unknown')
            dur = capability_duration_seconds(e, start_e)
            if dur is None:
                dur = 0
            # Get skill/tool name for this capability
            cap_inv_id = e.get('capabilityInvocationId', '')
            skill_name = cap_inv_names.get(cap_inv_id, '')
            step_events.append((e['timestamp'], 'capability', dur, '', [], cap_id, skill_name))

    step_events.sort(key=lambda x: x[0] if x[0] else datetime.min)
    for item in step_events:
        ts, typ = item[0], item[1]
        dur = item[2]
        if typ == 'model':
            fr, tools = item[3], item[4]
            tool_str = ','.join(tools) if tools else ''
            if fr == 'tool-calls' and tool_str:
                stage_timeline.append(f"Model({dur}s) → 调用 {tool_str}")
            elif fr == 'stop':
                stage_timeline.append(f"Model({dur}s) → 生成回答")
            else:
                stage_timeline.append(f"Model({dur}s) → {fr or '完成'}")
        else:
            cap_id, skill_name = item[5], item[6]
            label = f"{cap_id}" if not skill_name else f"{cap_id}({skill_name})"
            stage_timeline.append(f"{label}({dur}s)")

    # Session title
    title = ''
    for e in events:
        if e['event'] == 'session.title.generated':
            title = e['raw'].get('title', '')

    # Policy routing
    routing = ''
    for e in events:
        if e['event'] == 'policy.allowed':
            routing = e['details'].get('safeReasonCode', '') if e['details'] else ''

    # Matched RAG calls for this session (by time overlap)
    session_rag = [r for r in rag_matched
                   if session_start <= r['request_ts'] <= session_end]
    session_rag_indexes = []
    session_rag_index_count = 0
    for r in session_rag:
        session_rag_indexes.extend(r['indexes'])
        session_rag_index_count += r['indexCount']

    # Format multi-value fields
    sk_count = len(skill_durations)
    if sk_count == 1:
        sk_start_str = skill_start_times[0] if skill_start_times else ''
        sk_end_str = skill_end_times[0] if skill_end_times else ''
        sk_dur_str = skill_durations[0] if skill_durations else ''
    elif sk_count > 1:
        sk_start_str = '\n'.join(skill_start_times)
        sk_end_str = '\n'.join(skill_end_times)
        sk_dur_str = '\n'.join(str(d) for d in skill_durations)
    else:
        sk_start_str = ''
        sk_end_str = ''
        sk_dur_str = ''

    timeline_str = '\n'.join(stage_timeline)

    # RAG indexes detail per session
    rag_index_detail = ''
    if session_rag:
        idx_parts = []
        for r in session_rag:
            idx_str = ','.join(r['indexes']) if r['indexes'] else 'none'
            idx_parts.append(f"[{r['indexCount']}个索引] {idx_str} (耗时{r['duration_s']}s)")
        rag_index_detail = '\n'.join(idx_parts)

    # Query: prefer inputTextPreview, fallback to RAG query
    query_text = session_query.get(sid, '')
    # Match to test set for difficulty level
    matched_q, difficulty = match_question(query_text)
    # If no match, try matching via RAG query keywords
    if not difficulty and session_rag:
        import re as _re
        def _extract_kw(text):
            """Extract keywords from text - handles both Chinese and English, keyword-style and sentence-style."""
            # Split by common delimiters (Chinese/English punctuation, spaces)
            parts = _re.split(r'[，。？、；：！""（）\s,.;:!?()\-\d]+', text)
            return [p for p in parts if len(p) >= 2]
        def _extract_substrings(text, min_len=2, max_len=6):
            """Extract meaningful Chinese substrings for fuzzy matching."""
            # Remove punctuation, keep Chinese chars and alphanumeric
            clean = _re.sub(r'[，。？、；：！""（）\s,.;:!?()\-\d]+', '', text)
            result = set()
            for length in range(min_len, min(max_len + 1, len(clean) + 1)):
                for i in range(len(clean) - length + 1):
                    result.add(clean[i:i+length])
            return result

        for r in session_rag:
            rag_q = r.get('query', '').strip()
            if not rag_q:
                continue
            # 1. Exact or substring match
            for q in question_to_level:
                if rag_q in q or q in rag_q:
                    difficulty = question_to_level[q]
                    query_text = f'[RAG匹配] {rag_q}'
                    break
            if difficulty:
                break
            # 2. Keyword overlap match (>=2 overlapping keywords from delimiter split)
            kw_rag = set(_extract_kw(rag_q))
            if len(kw_rag) >= 2:
                best_match = None
                best_overlap = 0
                for q in question_to_level:
                    kw_q = set(_extract_kw(q))
                    overlap = len(kw_rag & kw_q)
                    if overlap > best_overlap:
                        best_overlap = overlap
                        best_match = q
                if best_overlap >= 2:
                    difficulty = question_to_level[best_match]
                    query_text = f'[RAG匹配] {rag_q}'
                    break
            # 3. RAG keywords contained in test question text (keyword-style RAG → sentence-style question)
            if len(kw_rag) >= 2:
                best_contain_match = None
                best_contain_count = 0
                for q in question_to_level:
                    contain_count = sum(1 for kw in kw_rag if kw in q)
                    if contain_count > best_contain_count:
                        best_contain_count = contain_count
                        best_contain_match = q
                if best_contain_count >= 2:
                    difficulty = question_to_level[best_contain_match]
                    query_text = f'[RAG匹配] {rag_q}'
                    break
            # 4. Substring containment: RAG query substrings in test question
            # For keyword-style RAG queries like "RRU5919支持的频段",
            # extract substrings and check if enough appear in any test question
            rag_subs = _extract_substrings(rag_q, min_len=2, max_len=8)
            if len(rag_subs) >= 3:
                best_sub_match = None
                best_sub_count = 0
                for q in question_to_level:
                    sub_count = sum(1 for sub in rag_subs if sub in q)
                    if sub_count > best_sub_count:
                        best_sub_count = sub_count
                        best_sub_match = q
                if best_sub_count >= 3:
                    difficulty = question_to_level[best_sub_match]
                    query_text = f'[RAG子串匹配] {rag_q}'
                    break

    all_rows.append({
        'sessionId': sid,
        'routing': routing,
        'skillCount': skill_count,
        'sessionStartTime': session_start.strftime('%Y-%m-%dT%H:%M:%S'),
        'sessionEndTime': session_end.strftime('%Y-%m-%dT%H:%M:%S'),
        'sessionDuration_s': session_duration,
        'skillStartTime': sk_start_str,
        'skillEndTime': sk_end_str,
        'skillDuration_s': sk_dur_str,
        'modelTotal_s': model_total,
        'modelCount': model_count,
        'capabilityTotal_s': cap_total,
        'capabilityCount': cap_count,
        'ragTotal_s': rag_total,
        'ragCount': rag_count,
        'ragIndexCount': session_rag_index_count,
        'bashTotal_s': round(cap_details.get('Bash', {}).get('total', 0), 2),
        'bashCount': cap_details.get('Bash', {}).get('count', 0),
        'readTotal_s': round(cap_details.get('Read', {}).get('total', 0), 2),
        'readCount': cap_details.get('Read', {}).get('count', 0),
        'writeTotal_s': round(cap_details.get('Write', {}).get('total', 0), 2),
        'writeCount': cap_details.get('Write', {}).get('count', 0),
        'pythonTotal_s': round(cap_details.get('Python', {}).get('total', 0), 2),
        'pythonCount': cap_details.get('Python', {}).get('count', 0),
        'otherDuration_s': other_duration,
        'stageTimeline': timeline_str,
        'stepDetail': '\n'.join(stage_timeline),
        'query': query_text,
        'difficulty': difficulty,
        'ragIndexDetail': rag_index_detail,
        'title': title,
    })

print(f"Total knowledge-qa sessions: {len(all_rows)}")

# ========== 5. Create Excel ==========
wb = Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def stats_rows(vals, label):
    if not vals:
        return [[f"{label}", "无数据"]]
    sv = sorted(vals)
    return [
        [f"{label} - 最小值", round(min(vals), 2)],
        [f"{label} - 最大值", round(max(vals), 2)],
        [f"{label} - 平均值", round(statistics.mean(vals), 2)],
        [f"{label} - 中位数", round(statistics.median(vals), 2)],
        [f"{label} - P90", round(sv[int(len(sv)*0.9)], 2)],
        [f"{label} - P95", round(sv[int(len(sv)*0.95)], 2)],
        [f"{label} - P99", round(sv[min(int(len(sv)*0.99), len(sv)-1)], 2)],
        [f"{label} - 样本数", len(vals)],
    ]

# ===== Sheet 1: 汇总统计 =====
ws1 = wb.active
ws1.title = "汇总统计"

durations = [r['sessionDuration_s'] for r in all_rows if r['sessionDuration_s'] >= 0]
skill_durations_flat = []
for r in all_rows:
    if isinstance(r['skillDuration_s'], (int, float)):
        skill_durations_flat.append(r['skillDuration_s'])
    elif isinstance(r['skillDuration_s'], str) and '\n' in r['skillDuration_s']:
        for part in r['skillDuration_s'].split('\n'):
            try:
                skill_durations_flat.append(float(part))
            except:
                pass
model_durations_list = [r['modelTotal_s'] for r in all_rows if r['modelTotal_s'] > 0]
rag_durations_list = [r['ragTotal_s'] for r in all_rows if r['ragTotal_s'] > 0]

summary_data = [
    ["指标", "值"],
    ["日志目录", LOG_DIR],
    ["分析时间范围(UTC)", f"{T_START.strftime('%Y-%m-%d %H:%M:%S')} ~ {T_END.strftime('%Y-%m-%d %H:%M:%S')}"],
    ["分析时间范围(本地+0800)", LOCAL_TIME_DESC],
    ["", ""],
    ["含Skill调用的会话数", len(kq_session_ids)],
    ["knowledge-qa总调用次数", sum(r['skillCount'] for r in all_rows)],
    ["", ""],
]
for stat in stats_rows(durations, "会话总耗时(秒)"):
    summary_data.append(stat)
summary_data.append(["", ""])
for stat in stats_rows(skill_durations_flat, "knowledge-qa调用耗时(秒)"):
    summary_data.append(stat)
summary_data.append(["", ""])
for stat in stats_rows(model_durations_list, "Model调用总耗时(秒)"):
    summary_data.append(stat)
summary_data.append(["", ""])
for stat in stats_rows(rag_durations_list, "Rag调用总耗时(秒)"):
    summary_data.append(stat)
if rag_matched:
    summary_data.append(["", ""])
    for stat in stats_rows([r['duration_s'] for r in rag_matched], "Rag单次检索耗时(秒)"):
        summary_data.append(stat)

# Index count stats
if index_counts:
    summary_data.append(["", ""])
    for stat in stats_rows(index_counts, "Rag每次请求索引数(indexes个数)"):
        summary_data.append(stat)

for row in summary_data:
    ws1.append(row)
for col in range(1, 3):
    ws1.cell(row=1, column=col).font = header_font
    ws1.cell(row=1, column=col).fill = header_fill
    ws1.cell(row=1, column=col).alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 45

# ===== Sheet 2: 会话明细 =====
ws2 = wb.create_sheet("会话明细")
headers = [
    "会话ID", "路由策略", "kq调用次数",
    "会话开始时间", "会话结束时间", "会话总耗时(秒)",
    "kq开始时间", "kq结束时间", "kq耗时(秒)",
    "Model总耗时(秒)", "Model调用次数",
    "能力调用总耗时(秒)", "能力调用次数",
    "Rag耗时(秒)", "Rag调用次数", "Rag索引总数",
    "Bash耗时(秒)", "Bash调用次数",
    "Read耗时(秒)", "Read调用次数",
    "Write耗时(秒)", "Write调用次数",
    "Python耗时(秒)", "Python调用次数",
    "其他耗时(秒)", "步骤详情(每步操作)", "阶段耗时明细", "查询问题(inputTextPreview)", "难度等级", "Rag索引明细", "会话标题",
]
ws2.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

for row_data in all_rows:
    values = [
        row_data['sessionId'], row_data['routing'], row_data['skillCount'],
        row_data['sessionStartTime'], row_data['sessionEndTime'], row_data['sessionDuration_s'],
        row_data['skillStartTime'], row_data['skillEndTime'], row_data['skillDuration_s'],
        row_data['modelTotal_s'], row_data['modelCount'],
        row_data['capabilityTotal_s'], row_data['capabilityCount'],
        row_data['ragTotal_s'], row_data['ragCount'], row_data['ragIndexCount'],
        row_data['bashTotal_s'], row_data['bashCount'],
        row_data['readTotal_s'], row_data['readCount'],
        row_data['writeTotal_s'], row_data['writeCount'],
        row_data['pythonTotal_s'], row_data['pythonCount'],
        row_data['otherDuration_s'], row_data['stepDetail'], row_data['stageTimeline'],
        row_data['query'], row_data['difficulty'], row_data['ragIndexDetail'], row_data['title'],
    ]
    ws2.append(values)
    row_num = ws2.max_row
    for col in range(1, len(values) + 1):
        cell = ws2.cell(row=row_num, column=col)
        cell.border = thin_border
        if col in (7, 8, 9, 26, 27, 28, 29, 30, 31):
            cell.alignment = Alignment(wrap_text=True, vertical='top')

col_widths = [42, 22, 10, 28, 28, 14, 28, 28, 16, 16, 10, 16, 10, 14, 10, 12, 14, 10, 14, 10, 14, 10, 14, 10, 12, 50, 50, 60, 12, 50, 40]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 3: 耗时分布 =====
ws3 = wb.create_sheet("耗时分布")

# Session duration distribution
ws3.append(["会话总耗时分布", "", ""])
for col in range(1, 4):
    ws3.cell(row=ws3.max_row, column=col).font = header_font
    ws3.cell(row=ws3.max_row, column=col).fill = header_fill
ws3.append(["耗时区间", "会话数", "占比"])
for col in range(1, 4):
    ws3.cell(row=ws3.max_row, column=col).font = Font(bold=True)

buckets = [
    (0, 10, "0-10s"), (10, 20, "10-20s"), (20, 30, "20-30s"),
    (30, 60, "30-60s"), (60, 90, "60-90s"), (90, 120, "90-120s"),
    (120, 180, "120-180s"), (180, 300, "180-300s"), (300, 9999, ">300s"),
]
for low, high, label in buckets:
    count = sum(1 for d in durations if low <= d < high)
    pct = f"{count/len(durations)*100:.1f}%" if durations else "0%"
    ws3.append([label, count, pct])

# knowledge-qa duration distribution
ws3.append([])
ws3.append(["knowledge-qa调用耗时分布", "", ""])
for col in range(1, 4):
    ws3.cell(row=ws3.max_row, column=col).font = header_font
    ws3.cell(row=ws3.max_row, column=col).fill = green_fill
ws3.append(["耗时区间", "调用次数", "占比"])
kq_buckets = [
    (0, 0.5, "0-0.5s"), (0.5, 1, "0.5-1s"), (1, 2, "1-2s"),
    (2, 5, "2-5s"), (5, 10, "5-10s"), (10, 30, "10-30s"), (30, 9999, ">30s"),
]
for low, high, label in kq_buckets:
    count = sum(1 for d in skill_durations_flat if low <= d < high)
    pct = f"{count/len(skill_durations_flat)*100:.1f}%" if skill_durations_flat else "0%"
    ws3.append([label, count, pct])

# Model per-invocation duration distribution
ws3.append([])
ws3.append(["Model调用耗时分布(每次调用)", "", ""])
for col in range(1, 4):
    ws3.cell(row=ws3.max_row, column=col).font = header_font
    ws3.cell(row=ws3.max_row, column=col).fill = orange_fill

all_model_inv_durations = []
for sid in kq_session_ids:
    events = sorted(session_events[sid], key=lambda e: e['timestamp'] or datetime.min)
    ms = {}
    for e in events:
        if e['event'] == 'model.invocation.started':
            ms[e['runId']] = e
        elif e['event'] == 'model.invocation.completed':
            se = ms.get(e['runId'])
            if se and se['timestamp'] and e['timestamp']:
                all_model_inv_durations.append(round((e['timestamp'] - se['timestamp']).total_seconds(), 3))
            elif e.get('durationMs') is not None:
                all_model_inv_durations.append(round(e['durationMs'] / 1000, 3))

ws3.append(["耗时区间", "调用次数", "占比"])
model_buckets = [
    (0, 3, "0-3s"), (3, 5, "3-5s"), (5, 10, "5-10s"),
    (10, 20, "10-20s"), (20, 30, "20-30s"), (30, 60, "30-60s"), (60, 9999, ">60s"),
]
for low, high, label in model_buckets:
    count = sum(1 for d in all_model_inv_durations if low <= d < high)
    pct = f"{count/len(all_model_inv_durations)*100:.1f}%" if all_model_inv_durations else "0%"
    ws3.append([label, count, pct])

# RAG duration distribution
if rag_matched:
    ws3.append([])
    ws3.append(["Rag检索耗时分布(每次调用)", "", ""])
    for col in range(1, 4):
        ws3.cell(row=ws3.max_row, column=col).font = header_font
        ws3.cell(row=ws3.max_row, column=col).fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    ws3.append(["耗时区间", "调用次数", "占比"])
    rag_durs = [r['duration_s'] for r in rag_matched]
    rag_buckets = [
        (0, 1, "0-1s"), (1, 2, "1-2s"), (2, 5, "2-5s"),
        (5, 10, "5-10s"), (10, 20, "10-20s"), (20, 30, "20-30s"), (30, 9999, ">30s"),
    ]
    for low, high, label in rag_buckets:
        count = sum(1 for d in rag_durs if low <= d < high)
        pct = f"{count/len(rag_durs)*100:.1f}%" if rag_durs else "0%"
        ws3.append([label, count, pct])

# RAG index count distribution
if index_counts:
    ws3.append([])
    ws3.append(["Rag每次请求索引数分布", "", ""])
    for col in range(1, 4):
        ws3.cell(row=ws3.max_row, column=col).font = header_font
        ws3.cell(row=ws3.max_row, column=col).fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ws3.append(["索引个数", "请求次数", "占比"])
    ic_counter = Counter(index_counts)
    for ic in sorted(ic_counter.keys()):
        cnt = ic_counter[ic]
        pct = f"{cnt/len(index_counts)*100:.1f}%"
        ws3.append([ic, cnt, pct])

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 10

# ===== Sheet 4: Rag检索明细 =====
ws4 = wb.create_sheet("Rag检索明细")
headers4 = ["请求时间", "响应时间", "耗时(秒)", "索引个数", "索引列表(indexes)", "查询内容", "topK", "HTTP状态", "响应体大小"]
ws4.append(headers4)
for col in range(1, len(headers4) + 1):
    ws4.cell(row=1, column=col).font = header_font
    ws4.cell(row=1, column=col).fill = header_fill
    ws4.cell(row=1, column=col).alignment = Alignment(horizontal='center')

for r in sorted(rag_matched, key=lambda x: x['request_ts'] or datetime.min):
    ws4.append([
        r['request_ts'].strftime('%Y-%m-%dT%H:%M:%S') if r['request_ts'] else '',
        r['response_ts'].strftime('%Y-%m-%dT%H:%M:%S') if r['response_ts'] else '',
        r['duration_s'],
        r['indexCount'],
        ', '.join(r['indexes']) if r['indexes'] else '',
        r['query'][:200],
        r['topK'],
        r['statusCode'],
        r['bodyLength'],
    ])
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 40
ws4.column_dimensions['F'].width = 60
ws4.column_dimensions['G'].width = 8
ws4.column_dimensions['H'].width = 10
ws4.column_dimensions['I'].width = 12

# ===== Sheet 5: Rag索引统计 =====
ws5 = wb.create_sheet("Rag索引统计")
ws5.append(["索引名称", "被请求次数", "占比"])
for col in range(1, 4):
    ws5.cell(row=1, column=col).font = header_font
    ws5.cell(row=1, column=col).fill = header_fill
    ws5.cell(row=1, column=col).alignment = Alignment(horizontal='center')

total_index_refs = sum(index_name_counter.values())
for idx_name, cnt in index_name_counter.most_common():
    pct = f"{cnt/total_index_refs*100:.1f}%" if total_index_refs else "0%"
    ws5.append([idx_name, cnt, pct])

# Index count vs duration correlation
ws5.append([])
ws5.append(["索引个数 vs 平均耗时", "", ""])
for col in range(1, 4):
    ws5.cell(row=ws5.max_row, column=col).font = header_font
    ws5.cell(row=ws5.max_row, column=col).fill = green_fill
ws5.append(["索引个数", "样本数", "平均耗时(秒)"])
ic_durations = defaultdict(list)
for r in rag_matched:
    ic_durations[r['indexCount']].append(r['duration_s'])
for ic in sorted(ic_durations.keys()):
    durs = ic_durations[ic]
    ws5.append([ic, len(durs), round(statistics.mean(durs), 3)])

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 15

# ===== Sheet 6: 并发统计 =====
ws6 = wb.create_sheet("并发统计")
minute_counts = defaultdict(int)
for r in all_rows:
    start = datetime.strptime(r['sessionStartTime'], '%Y-%m-%dT%H:%M:%S')
    end = datetime.strptime(r['sessionEndTime'], '%Y-%m-%dT%H:%M:%S')
    current = start.replace(second=0)
    end_minute = end.replace(second=0)
    while current <= end_minute:
        minute_counts[current.strftime('%Y-%m-%d %H:%M')] += 1
        current += timedelta(minutes=1)

ws6.append(["分钟", "活跃会话数"])
for minute in sorted(minute_counts.keys()):
    ws6.append([minute, minute_counts[minute]])
ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 12

# ===== Sheet 7: 按难度等级性能汇总 =====
ws7 = wb.create_sheet("难度等级性能汇总")

# Group rows by difficulty
level_rows = defaultdict(list)
for r in all_rows:
    lvl = r['difficulty'] if r['difficulty'] else '未匹配'
    level_rows[lvl].append(r)

# Collect per-session metrics for each level
level_metrics = {}
for lvl in sorted(level_rows.keys()):
    rows = level_rows[lvl]
    session_durs = [r['sessionDuration_s'] for r in rows if r['sessionDuration_s'] >= 0]
    kq_durs = []
    for r in rows:
        if isinstance(r['skillDuration_s'], (int, float)):
            kq_durs.append(r['skillDuration_s'])
        elif isinstance(r['skillDuration_s'], str) and '\n' in r['skillDuration_s']:
            for part in r['skillDuration_s'].split('\n'):
                try: kq_durs.append(float(part))
                except: pass
    model_durs = [r['modelTotal_s'] for r in rows if r['modelTotal_s'] > 0]
    rag_durs = [r['ragTotal_s'] for r in rows if r['ragTotal_s'] > 0]
    level_metrics[lvl] = {
        'count': len(rows),
        'session_durs': session_durs,
        'kq_durs': kq_durs,
        'model_durs': model_durs,
        'rag_durs': rag_durs,
    }

# Header
headers7 = [
    "难度等级", "会话数", "占比",
    "会话耗时-平均(秒)", "会话耗时-P50(秒)", "会话耗时-P90(秒)", "会话耗时-P95(秒)",
    "kq耗时-平均(秒)", "kq耗时-P50(秒)", "kq耗时-P90(秒)", "kq耗时-P95(秒)",
    "Model耗时-平均(秒)", "Model耗时-P50(秒)", "Model耗时-P90(秒)", "Model耗时-P95(秒)",
    "Rag耗时-平均(秒)", "Rag耗时-P50(秒)", "Rag耗时-P90(秒)", "Rag耗时-P95(秒)",
]
ws7.append(headers7)
for col in range(1, len(headers7) + 1):
    cell = ws7.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

total_sessions = len(all_rows)
# Sort: Level 1, 2, 3, 4, then 未匹配
level_order = []
for l in ['Level 1', 'Level 2', 'Level 3', 'Level 4', 'level 3', 'level 4']:
    if l in level_metrics:
        level_order.append(l)
if '未匹配' in level_metrics:
    level_order.append('未匹配')
# Add any remaining
for l in sorted(level_metrics.keys()):
    if l not in level_order:
        level_order.append(l)

for lvl in level_order:
    m = level_metrics[lvl]
    pct = f"{m['count']/total_sessions*100:.1f}%"
    def pct_val(vals, p):
        if not vals: return ''
        sv = sorted(vals)
        return round(sv[min(int(len(sv)*p), len(sv)-1)], 2)
    ws7.append([
        lvl, m['count'], pct,
        round(statistics.mean(m['session_durs']), 2) if m['session_durs'] else '',
        pct_val(m['session_durs'], 0.5), pct_val(m['session_durs'], 0.9), pct_val(m['session_durs'], 0.95),
        round(statistics.mean(m['kq_durs']), 2) if m['kq_durs'] else '',
        pct_val(m['kq_durs'], 0.5), pct_val(m['kq_durs'], 0.9), pct_val(m['kq_durs'], 0.95),
        round(statistics.mean(m['model_durs']), 2) if m['model_durs'] else '',
        pct_val(m['model_durs'], 0.5), pct_val(m['model_durs'], 0.9), pct_val(m['model_durs'], 0.95),
        round(statistics.mean(m['rag_durs']), 2) if m['rag_durs'] else '',
        pct_val(m['rag_durs'], 0.5), pct_val(m['rag_durs'], 0.9), pct_val(m['rag_durs'], 0.95),
    ])
    for col in range(1, len(headers7) + 1):
        ws7.cell(row=ws7.max_row, column=col).border = thin_border

ws7.column_dimensions['A'].width = 14
for i in range(2, len(headers7) + 1):
    ws7.column_dimensions[get_column_letter(i)].width = 16

# ===== Sheet 8: 各阶段耗时汇总 =====
ws8 = wb.create_sheet("各阶段耗时汇总")

# Collect all capability types and their durations across sessions
all_cap_durations = defaultdict(list)  # cap_id -> list of durations
all_model_durations = []  # per-invocation model durations

for sid in kq_session_ids:
    events = sorted(session_events[sid], key=lambda e: e['timestamp'] or datetime.min)
    cap_started_by_span_local = {}
    cap_started_pending_local = defaultdict(list)
    for e in events:
        if e['event'] == 'capability.started':
            remember_capability_start(cap_started_by_span_local, cap_started_pending_local, e)
        elif e['event'] == 'capability.completed':
            start_e = pop_capability_start(
                cap_started_by_span_local, cap_started_pending_local, e
            )
            cap_id = (e.get('details') or {}).get('capabilityId', 'Unknown')
            if cap_id == 'Unknown' and start_e:
                cap_id = (start_e.get('details') or {}).get('capabilityId', 'Unknown')
            dur = capability_duration_seconds(e, start_e)
            if dur is not None:
                all_cap_durations[cap_id].append(dur)

    # Model per-invocation
    model_started_local = {}
    for e in events:
        if e['event'] == 'model.invocation.started':
            model_started_local[e['runId']] = e
        elif e['event'] == 'model.invocation.completed':
            se = model_started_local.get(e['runId'])
            dur = None
            if se and se['timestamp'] and e['timestamp']:
                dur = round((e['timestamp'] - se['timestamp']).total_seconds(), 3)
            elif e.get('durationMs') is not None:
                dur = round(e['durationMs'] / 1000, 3)
            if dur is not None:
                all_model_durations.append(dur)

# Build summary table
def pct_val(vals, p):
    if not vals: return ''
    sv = sorted(vals)
    return round(sv[min(int(len(sv)*p), len(sv)-1)], 2)

headers8 = ["阶段(Capability)", "调用次数", "会话覆盖率",
            "耗时-平均(秒)", "耗时-P50(秒)", "耗时-P90(秒)", "耗时-P95(秒)", "耗时-P99(秒)", "耗时-最大(秒)"]
ws8.append(headers8)
for col in range(1, len(headers8) + 1):
    cell = ws8.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Add Model as a special "stage"
if all_model_durations:
    model_count = len(all_model_durations)
    sessions_with_model = sum(1 for sid in kq_session_ids
                              if any(e['event'] == 'model.invocation.completed'
                                     for e in session_events[sid]))
    ws8.append([
        'Model调用', model_count, f"{sessions_with_model}/{len(kq_session_ids)} ({sessions_with_model/len(kq_session_ids)*100:.1f}%)",
        round(statistics.mean(all_model_durations), 2),
        pct_val(all_model_durations, 0.5),
        pct_val(all_model_durations, 0.9),
        pct_val(all_model_durations, 0.95),
        pct_val(all_model_durations, 0.99),
        round(max(all_model_durations), 2),
    ])
    for col in range(1, len(headers8) + 1):
        ws8.cell(row=ws8.max_row, column=col).border = thin_border

# Add each capability
for cap_id in sorted(all_cap_durations.keys()):
    durs = all_cap_durations[cap_id]
    sessions_with_cap = sum(1 for sid in kq_session_ids
                            if any(e['event'] == 'capability.completed' and
                                   (e.get('details') or {}).get('capabilityId') == cap_id
                                   for e in session_events[sid]))
    ws8.append([
        cap_id, len(durs), f"{sessions_with_cap}/{len(kq_session_ids)} ({sessions_with_cap/len(kq_session_ids)*100:.1f}%)",
        round(statistics.mean(durs), 2),
        pct_val(durs, 0.5),
        pct_val(durs, 0.9),
        pct_val(durs, 0.95),
        pct_val(durs, 0.99),
        round(max(durs), 2),
    ])
    for col in range(1, len(headers8) + 1):
        ws8.cell(row=ws8.max_row, column=col).border = thin_border

ws8.column_dimensions['A'].width = 28
for i in range(2, len(headers8) + 1):
    ws8.column_dimensions[get_column_letter(i)].width = 16

# Save
_dir_name = os.path.basename(LOG_DIR.rstrip('/\\'))
_output_name = f'{_dir_name}_knowledge-qa性能分析_{TEST_SETS[TEST_SET_CONFIG]["name"]}.xlsx'
output_path = os.path.join(os.path.dirname(os.path.abspath(LOG_DIR)), _output_name)
# Fallback to LOG_DIR if parent path is locked
if os.path.exists(output_path):
    try:
        with open(output_path, 'a'):
            pass
    except PermissionError:
        output_path = os.path.join(LOG_DIR, _output_name)
wb.save(output_path)
print(f"\nExcel saved to: {output_path}")
print(f"Sheet 1: 汇总统计")
print(f"Sheet 2: 会话明细 ({len(all_rows)} rows)")
print(f"Sheet 3: 耗时分布")
print(f"Sheet 4: Rag检索明细 ({len(rag_matched)} rows)")
print(f"Sheet 5: Rag索引统计")
print(f"Sheet 6: 并发统计")
print(f"Sheet 7: 难度等级性能汇总 ({len(level_order)} levels)")
print(f"Sheet 8: 各阶段耗时汇总")
